from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List
import io
import openpyxl
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

router = APIRouter()

class TestCaseItem(BaseModel):
    title: str
    steps: str
    expected: str
    priority: str

class ExportRequest(BaseModel):
    test_cases: List[TestCaseItem]

@router.post("/export/excel")
async def export_excel(request: ExportRequest):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    headers = ["#", "Title", "Steps", "Expected Result", "Priority"]
    ws.append(headers)

    for col in range(1, 6):
        cell = ws.cell(row=1, column=col)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="4F81BD")

    for i, tc in enumerate(request.test_cases, start=1):
        ws.append([i, tc.title, tc.steps, tc.expected, tc.priority])

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 12

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=test_cases.xlsx"}
    )

@router.post("/export/pdf")
async def export_pdf(request: ExportRequest):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    elements = []

    # Independent style for table cells - black text, no shared reference
    cell_style = ParagraphStyle(
        'CellStyle',
        fontSize=8,
        leading=10,
        textColor=colors.black,
    )

    # Title
    elements.append(Paragraph("Test Cases Report", styles['Title']))
    elements.append(Spacer(1, 20))

    # Table data - wrap every cell in a Paragraph
    data = [["#", "Title", "Steps", "Expected", "Priority"]]
    for i, tc in enumerate(request.test_cases, start=1):
        data.append([
            Paragraph(str(i), cell_style),
            Paragraph(tc.title, cell_style),
            Paragraph(tc.steps, cell_style),
            Paragraph(tc.expected, cell_style),
            Paragraph(tc.priority, cell_style),
        ])

    table = Table(data, colWidths=[25, 90, 160, 150, 50], repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F81BD')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#DCE6F1')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=test_cases.pdf"}
    )